# -*- coding: utf-8 -*-
"""审计日志 Excel 导出内容回归（dual-mode-client Phase 1.3）。

pandas.DataFrame.to_excel 已替换为 openpyxl 直写；本测试锁定"写入→读回"
的内容契约：表头、列序、行数、单元格值与非字典项跳过行为。任何实现改动
（包括未来再换写入库）都必须保持该契约。
"""

from pathlib import Path

from openpyxl import load_workbook

from app.services.audit_service import AuditLogService
from app.services.audit_service_sync import AuditLogServiceSync


def _sample_logs():
    return [
        {
            "log_id": 1,
            "torrent_info_id": 100,
            "operation_type": "DELETE",
            "operator": "admin",
            "operation_time": "2026-08-23 10:00:00",
            "operation_result": "success",
            "error_message": "",
            "downloader_id": 3,
            "ip_address": "192.168.1.5",
            "user_agent": "pytest",
            "request_id": "req-1",
            "session_id": "sess-1",
            "operation_detail": "删除种子",
            "old_value": "on",
            "new_value": "off",
        },
        {
            "log_id": 2,
            "torrent_info_id": None,
            "operation_type": "LOGIN",
            "operator": "viewer",
            "operation_time": "2026-08-23 11:00:00",
            "operation_result": "failed",
            "error_message": "密码错误",
            "downloader_id": None,
            "ip_address": "10.0.0.8",
            "user_agent": "",
            "request_id": "req-2",
            "session_id": "",
            "operation_detail": "",
            "old_value": "",
            "new_value": "",
        },
    ]


class TestAsyncServiceExcelExport:
    async def test_content_contract(self, tmp_path: Path):
        service = AuditLogService(db_session=None)
        output = tmp_path / "audit.xlsx"
        assert await service.export_logs_to_excel(_sample_logs(), str(output)) is True

        wb = load_workbook(output)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 表头 + 2 数据行
        assert len(rows) == 3
        assert rows[0][0] == "日志ID"
        assert rows[0][3] == "操作人"
        assert rows[0][14] == "新值"
        # 首行数据按列序对齐
        assert rows[1][0] == 1
        assert rows[1][3] == "admin"
        assert rows[1][13] == "on"
        assert rows[1][14] == "off"
        # 第二行缺失键回退为空串
        assert rows[2][1] in (None, "")
        assert rows[2][6] == "密码错误"

    async def test_skips_non_dict_entries(self, tmp_path: Path, caplog):
        service = AuditLogService(db_session=None)
        output = tmp_path / "audit2.xlsx"
        logs = _sample_logs() + ["not-a-dict", None]
        assert await service.export_logs_to_excel(logs, str(output)) is True

        wb = load_workbook(output)
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 3  # 表头 + 2 条有效记录，非字典项被跳过

    async def test_empty_logs_writes_header_only(self, tmp_path: Path):
        service = AuditLogService(db_session=None)
        output = tmp_path / "audit3.xlsx"
        assert await service.export_logs_to_excel([], str(output)) is True
        rows = list(load_workbook(output).active.iter_rows(values_only=True))
        assert len(rows) == 1  # 仅表头


class TestSyncServiceExcelExport:
    def test_content_contract(self, tmp_path: Path):
        service = AuditLogServiceSync(db_session=None)
        output = tmp_path / "audit_sync.xlsx"
        logs = [
            {
                "id": 9,
                "torrent_info_id": 42,
                "operation_type": "EXPORT",
                "operator": "admin",
                "operation_time": "2026-08-23 12:00:00",
                "operation_result": "success",
                "error_message": None,
                "downloader_id": 1,
                "ip_address": "127.0.0.1",
                "request_id": "req-9",
                "session_id": "sess-9",
            }
        ]
        assert service.export_logs_to_excel(logs, str(output)) is True

        rows = list(load_workbook(output).active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[0][0] == "ID"
        assert rows[0][10] == "会话ID"
        assert rows[1][0] == 9
        assert rows[1][2] == "EXPORT"
        assert rows[1][10] == "sess-9"

    def test_bad_output_path_returns_false(self, tmp_path: Path):
        service = AuditLogServiceSync(db_session=None)
        # 目录作为输出路径必然写失败，函数必须吞异常返回 False
        assert service.export_logs_to_excel(_sample_logs(), str(tmp_path)) is False
